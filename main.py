import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook

File_Path = "data.xlsx"

if not os.path.exists(File_Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "UID"
    sheet["B1"] = "Password"
    workbook.save(File_Path)


def Login():
    UniqueID = identry.get()
    Password = Passwordentry.get()
    
    if not UniqueID or not Password:
        messagebox.showerror("Error", "Please fill all Entries")
        return
    
    workbook = load_workbook(File_Path)
    sheet = workbook.active
    
    for row in sheet.iter_rows(values_only=True):
        if row[0] == UniqueID and row[1] == Password:
            messagebox.showinfo("Welcome", "Successfully logged in")
            return
        
    messagebox.showerror("Error", "Wrong Info \nIf you don't have an account, Signup first")    

def Signup():
    UniqueID = identry.get()
    Password = Passwordentry.get()
    
    if not UniqueID or not Password:
        messagebox.showerror("Error", "Please fill all Entries")
        return
    
    workbook = load_workbook(File_Path)
    sheet = workbook.active
    
    for row in sheet.iter_rows(values_only=True):
        if row[0] == UniqueID:
            messagebox.showerror("Error", "Unique ID already exists \nPlease choose a diffrent ID")
            return
    
    sheet.append([UniqueID, Password])
    workbook.save(File_Path)
    messagebox.showinfo("Congratulations", "Account created successfully")
    identry.delete(0, tk.END)
    Passwordentry.delete(0, tk.END)
    

def change_to_signup():
    welcomeimglabel.place_forget()
    loginButton.place_forget()
    change_Page_Signup.place_forget()
    window.title("Signup")
    FirstLabel.config(text="Signup")
    idlabel.config(text="Choose a UID")
    Passwordlabel.config(text="Choose a Password")
    Helloimglabel.place(x=0, y=0)
    SignupButton.place(x=380, y=250, height=30, width=100)
    change_Page_login.place(x=355, y=300)
    identry.delete(0, tk.END)
    Passwordentry.delete(0, tk.END)
    

def change_to_login():
    Helloimglabel.place_forget()
    SignupButton.place_forget()
    change_Page_login.place_forget()
    window.title("Login")
    FirstLabel.config(text="Login")
    idlabel.config(text="Enter UID")
    Passwordlabel.config(text="Enter Password")
    welcomeimglabel.place(x=0, y=0)
    loginButton.place(x=380, y=250, height=30, width=100)
    change_Page_Signup.place(x=355, y=300)
    identry.delete(0, tk.END)
    Passwordentry.delete(0, tk.END)


window = tk.Tk()
window.title("Login")
window.geometry("600x400")
window.configure(bg="#e8e2d3")
window.resizable(False, False)

LeftFrame = tk.Frame(window)
LeftFrame.place(x=0, y=0, height=400, width=250)

welcomeimg = tk.PhotoImage(file="welcomeback.png")
welcomeimglabel = tk.Label(LeftFrame, image=welcomeimg)

Helloimg = tk.PhotoImage(file="HelloSir.png")
Helloimglabel = tk.Label(LeftFrame, image=Helloimg)

welcomeimglabel.place(x=0, y=0)

FirstLabel = tk.Label(window, text="Login", font=("Bold", 25), 
                      bg="#e8e2d3")
FirstLabel.place(x=380, y=50)

idlabel = tk.Label(window, text="Enter UID:", font=("Century", 10),
                   bg="#e8e2d3")
idlabel.place(x=320, y=110)

identry = tk.Entry(window, font=("Century", 12), bg="#c9c0a9", 
                   relief="sunken")
identry.place(x=350, y=130, height=30, width=200)

Passwordlabel = tk.Label(window, text="Enter Password:", font=("Century", 10),
                   bg="#e8e2d3")
Passwordlabel.place(x=320, y=180)

Passwordentry = tk.Entry(window, font=("Century", 15), bg="#c9c0a9", 
                   relief="sunken", show="*")
Passwordentry.place(x=350, y=200, height=30, width=200)

loginButton = tk.Button(window, text="Login", font=('bold',12),
                        relief="raised", bg="#5ce1e6", command=Login)

SignupButton = tk.Button(window, text="Signup", font=('bold',12),
                        relief="raised", bg="#5ce1e6", command=Signup)

loginButton.place(x=380, y=250, height=30, width=100)

change_Page_Signup = tk.Button(window, text="Don't have an account",font=('bold',12),
                        relief="flat", fg="#5ce1e6", bg="#e8e2d3", command=change_to_signup)

change_Page_login = tk.Button(window, text="Already have an account",font=('bold',12),
                        relief="flat", fg="#5ce1e6", bg="#e8e2d3", command=change_to_login)

change_Page_Signup.place(x=355, y=300)

window.mainloop()
